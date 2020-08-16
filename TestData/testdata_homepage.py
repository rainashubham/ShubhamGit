import openpyxl


class getdata_homepage:
    data =[{"firstname":"Shubham Raina","email":"Shubhamraina@gmail.com","password":"Password112"},{"firstname":"kalyani bodhankar","email":"kalyanibodhankar12@gmail.com","password":"PasswordXX1"}]

    @staticmethod
    def excel_utility(testcasename):
        sheet = openpyxl.load_workbook("C:\\Users\\pcw\\Desktop\\Book1.xlsx")

        Dict = {}

        worksheet = sheet.active

        for i in range(1, worksheet.max_row + 1):
            if worksheet.cell(row=i, column=1).value == testcasename:
                for j in range(2, worksheet.max_column + 1):
                    Dict[worksheet.cell(row=1, column=j).value] = worksheet.cell(row=i, column=j).value
                    print(worksheet.cell(row=i, column=j).value, end=",")
                print()

        return [Dict]