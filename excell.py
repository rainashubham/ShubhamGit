import openpyxl

sheet = openpyxl.load_workbook("C:\\Users\\pcw\\Desktop\\Book1.xlsx")

Dict = {}

worksheet = sheet.active

print(worksheet.max_row)

print(worksheet.max_column)
for i in range(1,worksheet.max_row+1):
    if worksheet.cell(row=i,column=1).value == "testcase2":
        for j in range(2,worksheet.max_column+1):
            Dict[worksheet.cell(row=1,column=j).value] = worksheet.cell(row=i,column=j).value
            print(worksheet.cell(row=i,column=j).value,end=",")
        print()

print(Dict)


